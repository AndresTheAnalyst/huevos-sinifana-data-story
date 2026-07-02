# -*- coding: utf-8 -*-
"""
construir_datos.py — Capa de datos del portafolio "Huevos Sinifana".

Lee las fuentes REALES del proyecto avícola (Huevos Sinifana SAS, Las Marías,
Caldas-Antioquia) y genera tablas limpias (tidy CSV, columnas en inglés) en data/,
listas para el dashboard, el notebook y el análisis. Nombres de columna en inglés
(estándar de portafolio internacional); las etiquetas para el usuario van bilingües
en el dashboard. Ver DATA_DICTIONARY.md.

Genera en data/:
  - production_weekly.csv   operativo: postura real vs estándar, huevos, alimento, mortalidad (semana × lote)
  - lot_comparison.csv      resumen: costo por huevo y margen por lote/época (con el fijo ya repartido)
  - egg_market_price.csv    mercado: precio mayorista huevo rojo Medellín, mensual (SIPSA)
  - feed_price.csv          insumo: precio del concentrado por mes (factura/estimado)
  - DATA_DICTIONARY.md      diccionario de datos bilingüe ES/EN

Honestidad: alimento con FACTURA real del proveedor donde existe (2020-03 → 2023-08);
antes, ESTIMADO con modelo maíz+soya calibrado contra facturas (R²=0.74). El costo fijo
se reparte entre la producción real de la granja día a día (lotes traslapados G3/G4
comparten nómina/servicios).
"""
from pathlib import Path
import csv
import openpyxl
import pandas as pd

AQUI = Path(__file__).resolve().parent
PORTAFOLIO = AQUI.parent
HARNESS = PORTAFOLIO.parent
AVICOLA = HARNESS / "agentes" / "director-avicola"

PROD_XLSX = AVICOLA / "produccion" / "resultados lotes actuales de produccion" / "Resultados_Produccion.xlsx"
SIPSA_XLSX = AVICOLA / "produccion" / "Precios_Huevo_Historico_SIPSA_2018_2026.xlsx"
CONC_CSV = AVICOLA / "analisis-costos" / "precio_concentrado_estimado.csv"
RESUMEN_CSV = AVICOLA / "facturas-alimento" / "resumen_precio_kg_mensual.csv"
COMP_XLSX = AVICOLA / "plantilla-avicultores" / "Comparativo_Total.xlsx"

DATA = PORTAFOLIO / "data"
DATA.mkdir(parents=True, exist_ok=True)

LOTES = ["G1", "G2", "G3", "G4", "G5", "G6"]

# Nombre canónico de la genética por lote (la fuente trae variantes: "Hy Line Brown" vs "Hy-Line Brown").
LINE_CANON = {
    "G1": "Hy-Line Brown", "G2": "Hy-Line Brown",
    "G3": "Babcock Brown", "G4": "Babcock Brown",
    "G5": "Lohmann Brown Classic", "G6": "Isa Brown",
}


def _iso(v):
    return v.date().isoformat() if hasattr(v, "year") else None


def build_production():
    """production_weekly.csv — una fila por semana y lote (datos operativos reales)."""
    wb = openpyxl.load_workbook(PROD_XLSX, data_only=True, read_only=True)
    out = []
    for lote in LOTES:
        ws = wb[lote]
        for r in range(6, ws.max_row + 1):
            sem = ws.cell(r, 1).value
            fi = ws.cell(r, 2).value
            if not (isinstance(sem, int) and hasattr(fi, "year")):
                continue
            if sem < 1:                      # descarta la fila artefacto "semana 0"
                continue
            out.append({
                "lot": lote,
                "line": LINE_CANON[lote],    # nombre canónico de la genética
                "week": sem,
                "date_start": _iso(fi),
                "date_end": _iso(ws.cell(r, 3).value),
                "hens": ws.cell(r, 4).value,
                "mortality": ws.cell(r, 5).value,
                "lay_rate_real": ws.cell(r, 6).value,        # % postura real
                "lay_rate_standard": ws.cell(r, 7).value,    # % postura estándar (tabla genética)
                "eggs": ws.cell(r, 9).value,
                "feed_kg": ws.cell(r, 11).value,
            })
    wb.close()
    df = pd.DataFrame(out)
    df.to_csv(DATA / "production_weekly.csv", index=False, encoding="utf-8")
    return df


def build_comparison():
    """lot_comparison.csv — resumen costo/margen por lote (lee el Comparativo_Total ya validado)."""
    wb = openpyxl.load_workbook(COMP_XLSX, data_only=True, read_only=True)
    ws = wb["Comparativo Total"]
    out = []
    for r in range(4, 10):                  # filas 4-9 = G1..G6 (encabezados en fila 3)
        lote = ws.cell(r, 1).value
        if not lote:
            continue
        margin = ws.cell(r, 11).value or 0
        out.append({
            "lot": lote,
            "line": LINE_CANON.get(lote, ws.cell(r, 2).value),
            "period": ws.cell(r, 3).value,
            "hens": ws.cell(r, 4).value,
            "lay_rate": ws.cell(r, 5).value,
            "cost_feed": ws.cell(r, 6).value,
            "cost_labor": ws.cell(r, 7).value,
            "cost_other": ws.cell(r, 8).value,
            "cost_total": ws.cell(r, 9).value,
            "sale_price": ws.cell(r, 10).value,
            "margin": margin,
            "status": "green" if margin > 30 else ("yellow" if margin >= 0 else "red"),
        })
    wb.close()
    df = pd.DataFrame(out)
    df.to_csv(DATA / "lot_comparison.csv", index=False, encoding="utf-8")
    return df


def build_market_price():
    """egg_market_price.csv — precio mayorista huevo rojo Medellín, mensual (SIPSA, ponderado por categoría)."""
    W = {"Huevo rojo extra": 0.12, "Huevo rojo AA": 0.49, "Huevo rojo A": 0.35, "Huevo rojo B": 0.04}
    dfp = pd.read_excel(SIPSA_XLSX, sheet_name="Base de Datos", header=3)
    dfp.columns = ["Fecha", "Ano", "Mes", "Producto", "Mercado", "Precio"]
    dfp["Precio"] = pd.to_numeric(dfp["Precio"], errors="coerce")
    mm = dfp[dfp["Mercado"].astype(str).str.contains("edell") & dfp["Producto"].isin(W)].dropna(subset=["Precio"]).copy()
    mm["month"] = mm["Fecha"].astype(str).str[:7]
    piv = mm.pivot_table(index="month", columns="Producto", values="Precio", aggfunc="mean")
    rows = []
    for mes, row in piv.iterrows():
        present = {t: row[t] for t in W if t in row and pd.notna(row[t])}
        avg = (sum(present[t] * W[t] for t in present) / sum(W[t] for t in present)) if present else None
        rows.append({
            "month": mes,
            "price_avg": round(avg) if avg else None,
            "price_extra": round(row["Huevo rojo extra"]) if "Huevo rojo extra" in row and pd.notna(row["Huevo rojo extra"]) else None,
            "price_aa": round(row["Huevo rojo AA"]) if "Huevo rojo AA" in row and pd.notna(row["Huevo rojo AA"]) else None,
            "price_a": round(row["Huevo rojo A"]) if "Huevo rojo A" in row and pd.notna(row["Huevo rojo A"]) else None,
            "price_b": round(row["Huevo rojo B"]) if "Huevo rojo B" in row and pd.notna(row["Huevo rojo B"]) else None,
        })
    df = pd.DataFrame(rows).sort_values("month")
    df.to_csv(DATA / "egg_market_price.csv", index=False, encoding="utf-8")
    return df


def build_feed_price():
    """feed_price.csv — precio del concentrado PONEDORA por mes. Estrategia honesta:
       · FACTURA real (del proveedor) donde exista: 2020-03 → 2023-08, ponderada por kg solo de productos PONEDORA.
       · si no hay factura, ESTIMADO por modelo maíz+soya (IndexMundi): 2016-05 → 2022-02.
       Real manda sobre estimado en los meses donde ambos existen."""
    # 1) estimado del modelo
    est = {}
    for row in csv.DictReader(open(CONC_CSV, encoding="utf-8-sig")):
        es = (row.get("precio_concentrado_ESTIMADO_cop_kg") or "").strip()
        if es:
            est[row["mes"]] = float(es)
    # 2) factura real: SOLO productos de ponedora (excluye pollitas/pollas y alimento de mascota)
    agg = {}
    for row in csv.DictReader(open(RESUMEN_CSV, encoding="utf-8-sig")):
        if "PONEDORA" in (row.get("producto") or "").upper():
            m = row["mes"]; kg = float(row.get("kg") or 0); tot = float(row.get("total_cop") or 0)
            a = agg.get(m, [0.0, 0.0]); a[0] += tot; a[1] += kg; agg[m] = a
    real = {m: v[0] / v[1] for m, v in agg.items() if v[1] > 0}
    # 3) combinar (real manda; si no, estimado)
    rows = []
    for m in sorted(set(est) | set(real)):
        if m in real:
            rows.append({"month": m, "price_per_kg": round(real[m]), "source": "invoice"})
        else:
            rows.append({"month": m, "price_per_kg": round(est[m]), "source": "estimated"})
    df = pd.DataFrame(rows).sort_values("month")
    df.to_csv(DATA / "feed_price.csv", index=False, encoding="utf-8")
    return df


def build_monthly(prod_df, feed_df, mkt_df):
    """lot_monthly.csv — economía MES A MES por lote: costo total/huevo, costo alimento/huevo,
    precio del productor y precio mayorista DANE-SIPSA. Mismo modelo de costo que el comparativo
    (nómina+servicios fijos repartidos entre la producción de la granja, sanidad por flock,
    empaque/calcio/ayudante variables), pero a resolución mensual."""
    smmlv = {int(r["anio"]): float(r["smmlv"])
             for r in csv.DictReader(open(AVICOLA / "analisis-costos" / "salario_minimo_colombia.csv",
                                          encoding="utf-8-sig"))}

    def fy(y):
        return smmlv.get(y, smmlv[2022]) / smmlv[2022]

    conc = dict(zip(feed_df["month"], feed_df["price_per_kg"]))
    src_map = dict(zip(feed_df["month"], feed_df["source"]))
    cms = sorted(conc)

    def cprice(m):                       # precio del concentrado (carry-forward si falta el mes)
        if m in conc:
            return conc[m]
        prev = [x for x in cms if x <= m]
        return conc[prev[-1]] if prev else conc[cms[0]]

    sipsa = dict(zip(mkt_df["month"], mkt_df["price_avg"]))
    FACTOR = 362 / 406                   # productor = mayorista × factor de finca (igual que el comparativo)
    NOMINA, SERVICIOS, SANIDAD = 3_942_313, 1_361_100, 50_000

    pw = prod_df[prod_df["eggs"] > 0].copy()
    pw["month"] = pw["date_start"].str[:7]
    g_all = pw.groupby(["lot", "month"], as_index=False).agg(eggs=("eggs", "sum"), feed_kg=("feed_kg", "sum"))
    farm_eggs = g_all.groupby("month")["eggs"].sum().to_dict()   # huevos REALES de toda la granja por mes (reparto del fijo)
    # descarta los meses-borde (arranque/cierre con pocos días) que disparan el costo/huevo a valores irreales
    peak = g_all.groupby("lot")["eggs"].transform("max")
    g = g_all[g_all["eggs"] >= 0.25 * peak].copy()
    # y el primer/último mes del lote si fue PARCIAL (ej. G6 salió el 23-oct-2022: ese mes carga la
    # nómina completa sobre medio mes de huevos → $/huevo irreal). Regla: borde con < 60% del mes vecino.
    drop = []
    for _, dl in g.groupby("lot"):
        dl = dl.sort_values("month")
        if len(dl) >= 2 and dl.iloc[0]["eggs"] < 0.6 * dl.iloc[1]["eggs"]:
            drop.append((dl.iloc[0]["lot"], dl.iloc[0]["month"]))
        if len(dl) >= 2 and dl.iloc[-1]["eggs"] < 0.6 * dl.iloc[-2]["eggs"]:
            drop.append((dl.iloc[-1]["lot"], dl.iloc[-1]["month"]))
    g = g[~g.apply(lambda r: (r["lot"], r["month"]) in drop, axis=1)].copy()

    rows = []
    for r in g.itertuples():
        y = int(r.month[:4])
        f = fy(y)
        feed_pe = r.feed_kg * cprice(r.month) / r.eggs                       # alimento/huevo (real del mes)
        shared = (NOMINA + SERVICIOS) * f / farm_eggs[r.month]               # nómina+servicios repartidos en la granja
        sanidad_pe = SANIDAD * f / r.eggs                                    # sanidad por flock (no se reparte)
        var_other = 230 * f / 30 + 3.4 * 274 * f / 1000 + 4 * f + 14000 * f / 370   # empaque+calcio+ayudante+otros
        total_pe = feed_pe + shared + sanidad_pe + var_other
        dane = sipsa.get(r.month)
        ok = dane is not None and pd.notna(dane)
        rows.append({
            "lot": r.lot, "month": r.month, "eggs": int(r.eggs),
            # precio del concentrado: real (factura) o estimado (modelo), según el mes; sin carry-forward
            "feed_price_kg": round(conc[r.month]) if r.month in conc else None,
            "feed_source": src_map.get(r.month),
            "feed_cost_egg": round(feed_pe),
            "total_cost_egg": round(total_pe),
            "producer_price": round(dane * FACTOR) if ok else None,
            "wholesale_dane": round(dane) if ok else None,
        })
    df = pd.DataFrame(rows).sort_values(["lot", "month"])
    df.to_csv(DATA / "lot_monthly.csv", index=False, encoding="utf-8")
    return df


DICTIONARY = """# Data Dictionary — Huevos Sinifana SAS

> Bilingual data dictionary (ES/EN). All monetary values in **COP** (Colombian pesos).
> Diccionario de datos bilingüe. Valores monetarios en **pesos colombianos (COP)**.

Source / Fuente: real records of **Huevos Sinifana SAS** (Las Marías farm, Caldas–Antioquia,
Colombia), 6 layer lots (G1–G6), 2016–2023. My own poultry operation / Operación avícola propia.

---

## `production_weekly.csv`
Weekly operating record, one row per lot × week. / Registro operativo semanal, una fila por lote × semana.

| Column | ES | EN |
|---|---|---|
| `lot` | Lote (G1–G6) | Lot id |
| `line` | Línea genética | Genetic line |
| `week` | Semana de vida del lote | Lot age week |
| `date_start` / `date_end` | Inicio / fin de la semana | Week start / end |
| `hens` | Aves vivas | Live hens |
| `mortality` | Mortalidad de la semana | Weekly mortality |
| `lay_rate_real` | % postura real | Actual lay rate |
| `lay_rate_standard` | % postura estándar (tabla genética) | Breed-standard lay rate |
| `eggs` | Huevos de la semana | Eggs that week |
| `feed_kg` | Alimento consumido (kg) | Feed consumed (kg) |

## `lot_comparison.csv`
Cost-per-egg and margin by lot/era. Fixed cost (labor+services) shared across overlapping lots. /
Costo por huevo y margen por lote/época. El costo fijo se reparte entre lotes traslapados.

| Column | ES | EN |
|---|---|---|
| `lot` / `line` / `period` | Lote / línea / época | Lot / line / era |
| `hens` | Aves (flock productivo) | Productive flock size |
| `lay_rate` | % postura promedio en plena postura | Avg peak-lay rate |
| `cost_feed` | Costo alimento / huevo | Feed cost / egg |
| `cost_labor` | Costo mano de obra / huevo (fijo repartido) | Labor cost / egg (shared fixed) |
| `cost_other` | Otros (servicios, sanidad, empaque, calcio) / huevo | Other costs / egg |
| `cost_total` | Costo total / huevo | Total cost / egg |
| `sale_price` | Precio de venta del productor / huevo | Producer sale price / egg |
| `margin` | Margen neto / huevo | Net margin / egg |
| `status` | Semáforo (green/yellow/red) | Traffic light |

**Honesty note / Nota de honestidad:** feed price from real supplier invoices where they
exist (2020-03 → 2023-08, covering G5–G6 fully and part of G3–G4); earlier months estimated
with a maize+soy model calibrated against invoices (R²=0.74). G1 sale price estimated (pre-SIPSA). /
Alimento con factura real donde existe (2020-03 → 2023-08); antes, estimado (R²=0,74).

## `egg_market_price.csv`
Monthly wholesale red-egg price, Medellín market (SIPSA-DANE). /
Precio mayorista mensual del huevo rojo, plaza Medellín (SIPSA-DANE).

| Column | ES | EN |
|---|---|---|
| `month` | Mes (YYYY-MM) | Month |
| `price_avg` | Precio ponderado (mezcla de categorías) | Weighted price |
| `price_extra` / `price_aa` / `price_a` / `price_b` | Precio por categoría | Price by grade |

## `feed_price.csv`
Monthly layer-concentrate price per kg. / Precio mensual del concentrado ponedora por kg.

| Column | ES | EN |
|---|---|---|
| `month` | Mes (YYYY-MM) | Month |
| `price_per_kg` | Precio por kg (COP) | Price per kg (COP) |
| `source` | `invoice` (factura real) / `estimated` (modelo) | Source |

## `lot_monthly.csv`
Month-by-month economics per lot (cost-per-egg model at monthly resolution). /
Economía mes a mes por lote (modelo de costo por huevo a resolución mensual).

| Column | ES | EN |
|---|---|---|
| `lot` / `month` | Lote / mes (YYYY-MM) | Lot / month |
| `eggs` | Huevos del mes | Eggs that month |
| `feed_price_kg` | Precio del concentrado ese mes (COP/kg) | Feed price that month (COP/kg) |
| `feed_source` | `invoice` (factura real) / `estimated` (modelo maíz+soya) | Source of the feed price |
| `feed_cost_egg` | Costo de alimento / huevo | Feed cost / egg |
| `total_cost_egg` | Costo total de producción / huevo | Total production cost / egg |
| `producer_price` | Precio de venta del productor / huevo (mayorista × factor de finca) | Producer sale price / egg |
| `wholesale_dane` | Precio mayorista DANE-SIPSA / huevo (Medellín) | DANE-SIPSA wholesale price / egg |

**Nota / Note:** "precio DANE" y "plaza mayorista" son la misma fuente (SIPSA, publicada por el DANE).
El precio del productor es ese mayorista × factor de finca (~0,89). Meses sin SIPSA (G1, pre-2018) → vacío.
Se excluyen los meses PARCIALES de entrada/salida del lote (cargan gastos fijos completos sobre pocos
días de producción y distorsionan el $/huevo). / Partial entry/exit months of each lot are excluded
(they load a full month of fixed costs onto a few days of production, distorting cost per egg).
"""


if __name__ == "__main__":
    prod = build_production()
    comp = build_comparison()
    mkt = build_market_price()
    feed = build_feed_price()
    monthly = build_monthly(prod, feed, mkt)
    (DATA / "DATA_DICTIONARY.md").write_text(DICTIONARY, encoding="utf-8")

    print("=" * 64)
    print("CAPA DE DATOS DEL PORTAFOLIO — generada en", DATA)
    print("=" * 64)
    print(f"  production_weekly.csv : {len(prod):>4} filas  ({prod['lot'].nunique()} lotes, "
          f"semanas {prod['week'].min()}-{prod['week'].max()})")
    print(f"  lot_comparison.csv    : {len(comp):>4} filas  (G1-G6)")
    print(f"  egg_market_price.csv  : {len(mkt):>4} filas  ({mkt['month'].min()} → {mkt['month'].max()})")
    print(f"  feed_price.csv        : {len(feed):>4} filas  ({feed['month'].min()} → {feed['month'].max()})")
    print(f"  lot_monthly.csv       : {len(monthly):>4} filas  (economía mes a mes por lote)")
    print(f"  DATA_DICTIONARY.md    : diccionario bilingüe")
    print()
    print("Comparativo de lotes (costo vs margen por huevo, COP):")
    print(comp[["lot", "line", "period", "cost_total", "sale_price", "margin", "status"]].to_string(index=False))
